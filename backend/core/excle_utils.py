import io
import logging
import os
from datetime import datetime, timezone

import openpyxl
import pytz
from openpyxl.styles import Font, Alignment

from core.aws_utils import download_image_from_local
from openpyxl.drawing.image import Image

from core.image_utils import plot_bounding_box


def create_excel_file(data_list, camera_name_dict):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Image Data"

    # Add headers
    headers = [
        "Location Name",
        "Camera Name",
        "Created Date",
        "Counts",
        "Label",
        "Image",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)  # Make headers bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row, document in enumerate(data_list, start=2):
        object_key = "/".join(document["image_url"].split("/")[3::])
        image_name = object_key.split("/")[-1]

        download_image_from_local(
            image_url=document["image_url"],
            save_path=image_name,
        )

        for data_detection in document["result"]["detection"]:
            plot_bounding_box(
                coordinate=data_detection["location"],
                img_name=image_name,
                label=data_detection["label"],
            )
        if int(document["camera_id"]) in camera_name_dict:
            sheet.cell(
                row=row, column=1, value=camera_name_dict[int(document["camera_id"])][1]
            ).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(
                row=row, column=2, value=camera_name_dict[int(document["camera_id"])][0]
            ).alignment = Alignment(horizontal="center", vertical="center")

        utc_datetime = datetime.fromtimestamp(
            int(document["created_date"]["$date"]) / 1000, tz=timezone.utc
        )
        ist_timezone = pytz.timezone("Asia/Kolkata")
        ist_datetime = utc_datetime.astimezone(ist_timezone)
        formatted_ist_time = ist_datetime.strftime("%B %d %Y, %I:%M:%S %p")
        sheet.cell(
            row=row,
            column=3,
            value=formatted_ist_time,
        ).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(
            row=row, column=4, value=sum(document["counts"].values())
        ).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(
            row=row, column=5, value=",".join(list(document["counts"].keys()))
        ).alignment = Alignment(horizontal="center", vertical="center")

        # Add image
        try:
            img = Image(image_name)
            img.width = 200  # Resize image (optional)
            img.height = 150
            cell_location = f"F{row}"
            sheet.add_image(img, cell_location)
            col_letter = cell_location[0]  # Extract column letter
            row_number = int(cell_location[1:])  # Extract row number
            sheet.column_dimensions[col_letter].width = img.width / 9
            sheet.row_dimensions[row_number].height = (img.height + 5) * 0.75
        except FileNotFoundError:
            sheet.cell(row=row, column=5, value="Image not found").alignment = (
                Alignment(horizontal="center", vertical="center")
            )
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        if col[0].value != "Image":
            for cell in col:
                try:
                    if cell.value:  # Check if cell has a value
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2  # Add padding
    output_file = io.BytesIO()
    wb.save(output_file)
    output_file.seek(0)
    for document in data_list:
        image_name = document["image_url"].split("/")[-1]
        if os.path.exists(image_name):
            os.remove(image_name)
    logging.info(f"Excel file created successfully!")
    return output_file
